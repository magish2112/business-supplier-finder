#!/usr/bin/env python3
"""Создаёт пример базы поставщиков: data/baza postavshiki.xlsx"""

from __future__ import annotations

import sys
from pathlib import Path

_ROOT = Path(__file__).resolve().parents[1]
if str(_ROOT) not in sys.path:
    sys.path.insert(0, str(_ROOT))


def main() -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    out_dir = _ROOT / "data"
    out_dir.mkdir(parents=True, exist_ok=True)
    out_path = out_dir / "baza postavshiki.xlsx"

    wb = Workbook()
    ws = wb.active
    ws.title = "Поставщики"

    headers = [
        "Наименование",
        "ИНН",
        "Город",
        "Направление деятельности",
        "Сайт",
        "Email",
        "Телефон",
        "Источник",
        "Статус проверки",
    ]
    head_font = Font(bold=True)
    head_fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = head_font
        c.fill = head_fill
        c.alignment = Alignment(horizontal="center", wrap_text=True)

    sample_rows = [
        [
            'ООО "СтройСнаб-М"',
            "7701234567",
            "Москва",
            "оптовые поставки сантехники",
            "https://example-santeh.ru",
            "opt@example-santeh.ru",
            "+7 (495) 000-11-22",
            "baza_postavshiki",
            "не проверен",
        ],
        [
            "ИП Петров А.А.",
            "",
            "Ставрополь",
            "металлопрокат опт",
            "",
            "metal@example.local",
            "+7 962 000-00-00",
            "baza_postavshiki",
            "",
        ],
        [
            'АО "ТрубПром"',
            "7707654321",
            "Санкт-Петербург",
            "трубы стальные и фитинги",
            "https://example-truby.spb.ru",
            "sales@example-truby.spb.ru",
            "+7 812 555-66-77",
            "baza_postavshiki",
            "inn_ok",
        ],
    ]
    for r, row in enumerate(sample_rows, 2):
        for col, val in enumerate(row, 1):
            ws.cell(row=r, column=col, value=val)

    for col in range(1, len(headers) + 1):
        letter = ws.cell(row=1, column=col).column_letter
        ws.column_dimensions[letter].width = min(28, max(12, len(headers[col - 1]) + 2))

    wb.save(out_path)
    print(f"Создан файл: {out_path}")


if __name__ == "__main__":
    main()
