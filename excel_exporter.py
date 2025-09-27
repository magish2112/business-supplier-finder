import pandas as pd
from datetime import datetime
from typing import List
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
from config import EXCEL_CONFIG, CATEGORIES
from models import Company

class ExcelExporter:
    """Экспортер данных в Excel"""
    
    def __init__(self):
        self.workbook = None
        self.filename = EXCEL_CONFIG['filename']
        
    def export_companies(self, companies: List[Company], filename: str = None):
        """Экспорт компаний в Excel с категоризацией"""
        if filename:
            self.filename = filename
            
        # Создаем новый workbook
        self.workbook = Workbook()
        
        # Удаляем дефолтный лист
        self.workbook.remove(self.workbook.active)
        
        # Создаем лист со всеми данными
        self._create_main_sheet(companies)
        
        # Создаем листы по категориям
        self._create_category_sheets(companies)
        
        # Создаем сводный лист
        self._create_summary_sheet(companies)
        
        # Сохраняем файл
        self.workbook.save(self.filename)
        print(f"Данные экспортированы в {self.filename}")
        
    def _create_main_sheet(self, companies: List[Company]):
        """Создание основного листа со всеми данными"""
        ws = self.workbook.create_sheet("Все компании")
        
        # Заголовки
        headers = EXCEL_CONFIG['columns'] + ['Источник']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
            
        # Данные
        for row, company in enumerate(companies, 2):
            data = company.to_dict()
            for col, header in enumerate(headers, 1):
                ws.cell(row=row, column=col, value=data.get(header, ''))
                
        # Автоматическая ширина столбцов
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
            
    def _create_category_sheets(self, companies: List[Company]):
        """Создание листов по категориям"""
        # Группируем компании по категориям
        companies_by_category = {}
        for company in companies:
            category = company.category
            if category not in companies_by_category:
                companies_by_category[category] = []
            companies_by_category[category].append(company)
            
        # Создаем лист для каждой категории
        for category, category_companies in companies_by_category.items():
            ws = self.workbook.create_sheet(category)
            
            # Заголовки
            headers = EXCEL_CONFIG['columns'] + ['Источник']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="E6F3FF", end_color="E6F3FF", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")
                
            # Данные
            for row, company in enumerate(category_companies, 2):
                data = company.to_dict()
                for col, header in enumerate(headers, 1):
                    cell = ws.cell(row=row, column=col, value=data.get(header, ''))
                    
                    # Цветовая индикация статуса
                    if header == 'Статус':
                        if data.get(header) == 'Работает':
                            cell.fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
                        elif data.get(header) == 'Не работает':
                            cell.fill = PatternFill(start_color="FFB6C1", end_color="FFB6C1", fill_type="solid")
                            
            # Автоматическая ширина столбцов
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
                
    def _create_summary_sheet(self, companies: List[Company]):
        """Создание сводного листа"""
        ws = self.workbook.create_sheet("Сводка")
        
        # Статистика по категориям
        category_stats = {}
        status_stats = {'Работает': 0, 'Не работает': 0, 'Неизвестно': 0}
        
        for company in companies:
            # По категориям
            category = company.category
            if category not in category_stats:
                category_stats[category] = {'total': 0, 'active': 0, 'inactive': 0}
            category_stats[category]['total'] += 1
            
            if company.status.value == 'Работает':
                category_stats[category]['active'] += 1
            elif company.status.value == 'Не работает':
                category_stats[category]['inactive'] += 1
                
            # По статусам
            status_stats[company.status.value] += 1
            
        # Заголовок
        ws['A1'] = "Сводка по собранным данным"
        ws['A1'].font = Font(bold=True, size=16)
        
        # Общая статистика
        ws['A3'] = "Общая статистика:"
        ws['A3'].font = Font(bold=True)
        
        row = 4
        ws[f'A{row}'] = f"Всего компаний: {len(companies)}"
        row += 1
        ws[f'A{row}'] = f"Работающих: {status_stats['Работает']}"
        row += 1
        ws[f'A{row}'] = f"Не работающих: {status_stats['Не работает']}"
        row += 1
        ws[f'A{row}'] = f"Неизвестно: {status_stats['Неизвестно']}"
        row += 1
        
        # Статистика по категориям
        row += 2
        ws[f'A{row}'] = "Статистика по категориям:"
        ws[f'A{row}'].font = Font(bold=True)
        
        row += 1
        headers = ['Категория', 'Всего', 'Работает', 'Не работает', 'Процент работающих']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            
        row += 1
        for category, stats in category_stats.items():
            ws.cell(row=row, column=1, value=category)
            ws.cell(row=row, column=2, value=stats['total'])
            ws.cell(row=row, column=3, value=stats['active'])
            ws.cell(row=row, column=4, value=stats['inactive'])
            
            # Процент работающих
            if stats['total'] > 0:
                percentage = (stats['active'] / stats['total']) * 100
                ws.cell(row=row, column=5, value=f"{percentage:.1f}%")
            else:
                ws.cell(row=row, column=5, value="0%")
            row += 1
            
        # Автоматическая ширина столбцов
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
            
    def export_filtered(self, companies: List[Company], category: str = None, status: str = None):
        """Экспорт отфильтрованных данных"""
        filtered_companies = companies
        
        if category:
            filtered_companies = [c for c in filtered_companies if c.category == category]
            
        if status:
            filtered_companies = [c for c in filtered_companies if c.status.value == status]
            
        filename = f"filtered_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        self.export_companies(filtered_companies, filename)
        
        return filename 