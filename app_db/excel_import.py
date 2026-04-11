"""
Импорт поставщиков из Excel (пример: «baza postavshiki») в таблицу suppliers.

Первая строка — заголовки. Поддерживаются русские и английские названия колонок.
"""

from __future__ import annotations

import re
import uuid
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

from app_db.repositories import SupplierRepository

# Нормализованный заголовок → ключ поля БД
HEADER_ALIASES: Dict[str, str] = {
    "наименование": "name",
    "название": "name",
    "название компании": "name",
    "name": "name",
    "компания": "name",
    "инн": "inn",
    "inn": "inn",
    "город": "city",
    "city": "city",
    "регион": "city",
    "направление деятельности": "activity_direction",
    "направление": "activity_direction",
    "вид деятельности": "activity_direction",
    "категория": "activity_direction",
    "activity_direction": "activity_direction",
    "сайт": "website_url",
    "вебсайт": "website_url",
    "веб-сайт": "website_url",
    "url": "website_url",
    "website": "website_url",
    "email": "email",
    "почта": "email",
    "e-mail": "email",
    "телефон": "phone",
    "phone": "phone",
    "тел": "phone",
    "источник": "source",
    "source": "source",
    "статус проверки": "verification_status",
    "verification_status": "verification_status",
    "статус": "verification_status",
    "примечания": "notes_json",
    "notes": "notes_json",
}


def _norm_header(h: Any) -> str:
    if h is None:
        return ""
    t = str(h).strip().lower()
    t = re.sub(r"\s+", " ", t)
    return t


def _resolve_excel_path(path: Path, project_root: Optional[Path] = None) -> Path:
    p = path.expanduser()
    if p.is_absolute():
        return p
    root = project_root or Path(__file__).resolve().parents[1]
    return (root / p).resolve()


def import_suppliers_from_excel(
    excel_path: Path,
    *,
    db_path: Optional[str] = None,
    sheet_name: Optional[str] = None,
    project_root: Optional[Path] = None,
    skip_empty_name: bool = True,
) -> Tuple[int, int]:
    """
    Читает лист и делает insert_or_update по каждой строке данных.

    Возвращает (импортировано_строк, пропущено_пустых_имён).
    """
    try:
        from openpyxl import load_workbook  # type: ignore
    except ImportError as e:
        raise RuntimeError("Установите openpyxl: pip install openpyxl") from e

    xlsx = _resolve_excel_path(Path(excel_path), project_root)
    if not xlsx.is_file():
        raise FileNotFoundError(f"Файл Excel не найден: {xlsx}")

    wb = load_workbook(xlsx, read_only=True, data_only=True)
    if sheet_name:
        ws = wb[sheet_name]
    else:
        ws = wb.active

    rows = ws.iter_rows(values_only=True)
    try:
        header_row = next(rows)
    except StopIteration:
        wb.close()
        return 0, 0

    col_index: Dict[str, int] = {}
    for i, cell in enumerate(header_row):
        key = HEADER_ALIASES.get(_norm_header(cell))
        if key and key not in col_index:
            col_index[key] = i

    if "name" not in col_index:
        wb.close()
        raise ValueError(
            "В первой строке нужна колонка с названием организации "
            "(например: «Наименование», «Название», «Название компании»)."
        )

    imported = 0
    skipped = 0
    with SupplierRepository(db_path=db_path) as repo:
        for row in rows:
            if row is None:
                continue
            cells: List[Any] = list(row)
            def get(field: str) -> Optional[str]:
                idx = col_index.get(field)
                if idx is None or idx >= len(cells):
                    return None
                v = cells[idx]
                if v is None:
                    return None
                s = str(v).strip()
                return s if s else None

            name = get("name")
            if not name:
                if skip_empty_name:
                    skipped += 1
                continue

            notes = get("notes_json")
            data: Dict[str, Any] = {
                "name": name,
                "inn": get("inn"),
                "city": get("city"),
                "activity_direction": get("activity_direction"),
                "website_url": get("website_url"),
                "email": get("email"),
                "phone": get("phone"),
                "source": get("source") or "excel",
                "verification_status": get("verification_status"),
                "notes_json": notes,
            }
            repo.insert_or_update(data)
            imported += 1

    wb.close()
    return imported, skipped
